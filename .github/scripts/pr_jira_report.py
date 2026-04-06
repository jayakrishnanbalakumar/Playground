"""Generate an Excel report that maps repository PRs to Jira ticket keys."""

import json
import os
import re
import urllib.parse
import urllib.request
from datetime import datetime, timedelta, timezone
from typing import Dict, Iterable, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

GITHUB_API = "https://api.github.com"
JIRA_KEY_PATTERN = re.compile(r"\bJPM-\d+\b")


def parse_link_header(link_header: str) -> Optional[str]:
    """Return the URL for the next page from a GitHub Link header."""
    if not link_header:
        return None

    parts = [p.strip() for p in link_header.split(",")]
    for part in parts:
        if 'rel="next"' in part:
            left = part.find("<")
            right = part.find(">")
            if left != -1 and right != -1:
                return part[left + 1 : right]
    return None


def github_get(url: str, token: str) -> Tuple[List[Dict], Optional[str]]:
    """Fetch one GitHub API page and return JSON data with the next page URL."""
    headers = {
        "Accept": "application/vnd.github+json",
        "Authorization": f"Bearer {token}",
        "User-Agent": "daily-pr-jira-report",
        "X-GitHub-Api-Version": "2022-11-28",
    }

    request = urllib.request.Request(url=url, headers=headers, method="GET")
    with urllib.request.urlopen(request) as response:
        data = json.loads(response.read().decode("utf-8"))
        next_url = parse_link_header(response.headers.get("Link", ""))
        return data, next_url


def paginate(url: str, token: str) -> List[Dict]:
    """Follow paginated GitHub API responses and collect all items."""
    all_items: List[Dict] = []
    next_url: Optional[str] = url
    while next_url:
        items, next_url = github_get(next_url, token)
        all_items.extend(items)
    return all_items


def extract_jira_keys(text: str) -> Set[str]:
    """Extract unique Jira-style keys from free-form text."""
    if not text:
        return set()
    return set(JIRA_KEY_PATTERN.findall(text))


def build_ticket_text(keys: Iterable[str]) -> str:
    """Build the Excel cell text for Jira ticket references."""
    unique_sorted = sorted(set(keys))
    if not unique_sorted:
        return "-"
    return ", ".join(unique_sorted)


def set_column_widths(worksheet) -> None:
    """Set workbook column widths based on populated cell values."""
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def with_query(url: str, params: Dict[str, str]) -> str:
    """Append URL-encoded query parameters to a base URL."""
    return f"{url}?{urllib.parse.urlencode(params)}"


def parse_input_date(date_text: str, name: str) -> datetime:
    """Parse a YYYY-MM-DD input date into a UTC datetime at midnight."""
    try:
        return datetime.strptime(date_text, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except ValueError as exc:
        raise RuntimeError(f"{name} must be in YYYY-MM-DD format") from exc


def parse_github_datetime(value: str) -> datetime:
    """Parse GitHub ISO-8601 timestamps such as 2026-03-27T10:20:30Z."""
    return datetime.fromisoformat(value.replace("Z", "+00:00")).astimezone(timezone.utc)


def resolve_date_range(
    from_date_text: str,
    to_date_text: str,
    now_utc: datetime,
) -> Tuple[datetime, datetime]:
    """Resolve input dates, defaulting to Jan 1 through today when omitted."""
    start_of_year = datetime(now_utc.year, 1, 1, tzinfo=timezone.utc)
    start_date = parse_input_date(from_date_text, "FROM_DATE") if from_date_text else start_of_year
    end_date_inclusive = (
        parse_input_date(to_date_text, "TO_DATE") if to_date_text else now_utc.replace(hour=0, minute=0, second=0, microsecond=0)
    )
    end_date_exclusive = end_date_inclusive + timedelta(days=1)

    if start_date >= end_date_exclusive:
        raise RuntimeError("FROM_DATE must be earlier than or equal to TO_DATE")

    return start_date, end_date_exclusive


def is_within_date_range(
    updated_at: str,
    from_date: Optional[datetime],
    to_date_exclusive: Optional[datetime],
) -> bool:
    """Return True when updated_at falls within optional date boundaries."""
    if not updated_at:
        return False

    updated_dt = parse_github_datetime(updated_at)
    if from_date and updated_dt < from_date:
        return False
    if to_date_exclusive and updated_dt >= to_date_exclusive:
        return False
    return True


def collect_pr_jira_keys(
    owner: str,
    repo: str,
    pr_number: int,
    token: str,
    pr_title: str,
    pr_body: str,
) -> Set[str]:
    """Collect Jira keys from a PR title/body and its issue and review comments."""
    keys = set()
    keys |= extract_jira_keys(pr_title)
    keys |= extract_jira_keys(pr_body)

    issue_comments_url = with_query(
        f"{GITHUB_API}/repos/{owner}/{repo}/issues/{pr_number}/comments",
        {"per_page": "100"},
    )
    issue_comments = paginate(issue_comments_url, token)
    for comment in issue_comments:
        keys |= extract_jira_keys(comment.get("body", ""))

    review_comments_url = with_query(
        f"{GITHUB_API}/repos/{owner}/{repo}/pulls/{pr_number}/comments",
        {"per_page": "100"},
    )
    review_comments = paginate(review_comments_url, token)
    for comment in review_comments:
        keys |= extract_jira_keys(comment.get("body", ""))

    return keys


def main() -> None:
    """Generate the PR-to-Jira Excel report and write it to reports/."""
    now_utc = datetime.now(timezone.utc)
    token = os.environ.get("GITHUB_TOKEN", "").strip()
    repo_full_name = os.environ.get("REPO", "").strip()
    jira_base_url = os.environ.get("JIRA_BASE_URL", "").strip()
    pr_base_branch = os.environ.get("PR_BASE_BRANCH", "").strip()
    from_date_text = os.environ.get("FROM_DATE", "").strip()
    to_date_text = os.environ.get("TO_DATE", "").strip()

    if not token:
        raise RuntimeError("GITHUB_TOKEN is required")
    if "/" not in repo_full_name:
        raise RuntimeError("REPO must look like owner/repository")

    owner, repo = repo_full_name.split("/", 1)
    from_date, to_date_exclusive = resolve_date_range(from_date_text, to_date_text, now_utc)

    pulls_url = with_query(
        f"{GITHUB_API}/repos/{owner}/{repo}/pulls",
        {
            "state": "all",
            "sort": "updated",
            "direction": "desc",
            "per_page": "100",
        },
    )
    if pr_base_branch:
        pulls_url = with_query(
            f"{GITHUB_API}/repos/{owner}/{repo}/pulls",
            {
                "state": "all",
                "sort": "updated",
                "direction": "desc",
                "per_page": "100",
                "base": pr_base_branch,
            },
        )
    pulls = paginate(pulls_url, token)
    pulls = [
        pr
        for pr in pulls
        if is_within_date_range(pr.get("updated_at") or "", from_date, to_date_exclusive)
    ]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PR Jira Report"
    worksheet.append(["Generated (UTC)", now_utc.isoformat()])
    worksheet.append([])
    headers = ["PR", "Title", "State", "Jira Tickets", "Updated"]
    worksheet.append(headers)

    for cell in worksheet[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")

    for pr in pulls:
        pr_number = pr.get("number")
        pr_title = pr.get("title") or ""
        title = pr_title.replace("|", " ").strip()
        pr_url = pr.get("html_url") or ""
        pr_body = pr.get("body") or ""

        state = pr.get("state", "unknown")
        if pr.get("merged_at"):
            state = "merged"

        updated_at = pr.get("updated_at") or "-"

        keys = collect_pr_jira_keys(owner, repo, int(pr_number), token, pr_title, pr_body)
        ticket_cell = build_ticket_text(keys)

        worksheet.append([f"#{pr_number}", title, state, ticket_cell, updated_at])
        row_index = worksheet.max_row

        pr_cell = worksheet.cell(row=row_index, column=1)
        if pr_url:
            pr_cell.hyperlink = pr_url
            pr_cell.style = "Hyperlink"

        jira_cell = worksheet.cell(row=row_index, column=4)
        clean_base = jira_base_url.strip().rstrip("/")
        if clean_base and len(keys) == 1:
            jira_key = next(iter(keys))
            jira_cell.hyperlink = f"{clean_base}/browse/{jira_key}"
            jira_cell.style = "Hyperlink"

        for column_index in range(1, 6):
            worksheet.cell(row=row_index, column=column_index).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    worksheet.freeze_panes = "A4"
    worksheet.auto_filter.ref = f"A3:E{worksheet.max_row}"
    set_column_widths(worksheet)

    os.makedirs("reports", exist_ok=True)
    output_path = os.path.join("reports", "pr-jira-report.xlsx")
    workbook.save(output_path)


if __name__ == "__main__":
    main()
